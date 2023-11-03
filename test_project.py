import pytest
import os
import openpyxl
import warnings
from unittest.mock import patch
from openpyxl import Workbook
from project import get_input_data, request_output, request_price, calculator
from create_data_file import create_xls
from create_report import create_report
from user_data import User


# Test created input data file
def test_created_data_file_contents():
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    create_xls()  # Generate input data file
    assert os.path.exists("Data.xlsx")
    wb = openpyxl.load_workbook("Data.xlsx")
    ws = wb.active
    # Check if specific cells in the generated Excel file contain the expected values
    assert ws["A1"].value == "Input Data for Solar Panel Calculator"
    assert ws["A3"].value == "Address"
    assert ws["A11"].value == "Azimuth of rooftop"


# Test the initialization of the User class
def test_user_initialization():
    user = User(
        "7415 Southwest Pkwy Building 8",
        "Austin",
        "TX",
        "United States",
        100,
        5000,
        "Standard",
        "Fixed(open track)",
        180,
    )
    assert user.address == "7415 Southwest Pkwy Building 8"
    assert user.city == "Austin"
    assert user.state == "TX"
    assert user.country == "United States"
    assert user.rooftop_size == 100
    assert user.annual_consumption == 5000
    assert user.module_type == "Standard"
    assert user.array_type == "Fixed(open track)"
    assert user.azimuth == 180


# Test the getter and setter methods
def test_user_getters_and_setters():
    user = User(
        "7415 Southwest Pkwy Building 8",
        "Austin",
        "TX",
        "United States",
        100,
        5000,
        "Standard",
        "Fixed(open track)",
        180,
    )
    user.annual_output = 6000
    user.price = 15000
    user.monthly_output = [500, 600, 700]
    assert user.annual_output == 6000
    assert user.price == 15000
    assert user.monthly_output == [500, 600, 700]


# Test the get_module_key and get_array_key methods
def test_get_module_key_and_array_key():
    user = User(
        "7415 Southwest Pkwy Building 8",
        "Austin",
        "TX",
        "United States",
        100,
        5000,
        "Standard",
        "Fixed(open track)",
        180,
    )
    assert user.get_module_key() == "0"
    assert user.get_array_key() == "0"


# Test created report file
def test_created_report_file_contents():
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    user = User(
        "7415 Southwest Pkwy Building 8",
        "Austin",
        "TX",
        "United States",
        100,
        5000,
        "Standard",
        "Fixed(open track)",
        180,
    )
    report = [11.5, 55.4, 28.8, 17277, 1325, 13.1]
    create_report(user, report)  # Generate report file
    wb = openpyxl.load_workbook("Report.xlsx")
    ws = wb.active
    # Check if specific cells in the generated Excel file contain the expected values
    assert ws["A1"].value == "Solar Panel Report"
    assert ws["A3"].value == "Annual electricity consumption"
    assert ws["A11"].value == "Total repayment period"
    assert ws["D4"].value == "March"
    assert ws["B3"].value == user.annual_consumption
    assert ws["B11"].value == 13.1


# Test project.py code
# Create a sample Excel file for testing
@pytest.fixture
def sample_excel_file(tmp_path):
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    file_path = tmp_path / "sample_data.xlsx"
    wb = Workbook()
    ws = wb.active
    data = [
        "7415 Southwest Pkwy Building 8",
        "Austin",
        "TX",
        "USA",
        100,
        5000,
        "Standard",
        "Fixed(roof mount)",
        180,
    ]
    for i, value in enumerate(data, start=3):
        ws.cell(row=i, column=2).value = value
    wb.save(file_path)
    return file_path


# Test the get_input_data function
def test_get_input_data(sample_excel_file):
    user = get_input_data(sample_excel_file)
    assert user.address == "7415 Southwest Pkwy Building 8"
    assert user.city == "Austin"
    assert user.state == "TX"
    assert user.rooftop_size == 100
    assert user.module_type == "Standard"
    assert user.azimuth == 180


# Mocking the requests.get function
@patch("requests.get")
def test_request_output(mock_get):
    mock_response = {
        "outputs": {
            "ac_annual": 5000,
            "ac_monthly": [
                400,
                350,
                450,
                600,
                700,
                800,
                900,
                1000,
                1100,
                1200,
                1300,
                1400,
            ],
        }
    }
    mock_get.return_value.json.return_value = mock_response

    user = User(
        "7415 Southwest Pkwy Building 8",
        "Austin",
        "TX",
        "United States",
        100,
        5000,
        "Standard",
        "Fixed(open track)",
        180,
    )
    request_output(user, 1000)

    assert user.annual_output == 5000
    assert user.monthly_output == [
        400,
        350,
        450,
        600,
        700,
        800,
        900,
        1000,
        1100,
        1200,
        1300,
        1400,
    ]


# Mocking the requests.get function
@patch("requests.get")
def test_request_price(mock_get):
    mock_response = {"outputs": {"residential": 0.12}}
    mock_get.return_value.json.return_value = mock_response

    user = User(
        "7415 Southwest Pkwy Building 8",
        "Austin",
        "TX",
        "United States",
        100,
        5000,
        "Standard",
        "Fixed(open track)",
        180,
    )
    request_price(user)

    assert user.price == 0.12


# Test the calculator function
def test_calculator():
    user = User(
        "7415 Southwest Pkwy Building 8",
        "Austin",
        "TX",
        "United States",
        100,
        5000,
        "Standard",
        "Fixed(open track)",
        180,
    )
    report = calculator(user)
    assert round(report[0], 2) == 5.24
    assert len(report) == 6
