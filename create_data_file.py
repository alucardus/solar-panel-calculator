from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.worksheet.datavalidation import DataValidation


def main():
    create_xls()

# Creating Excel file for input data
def create_xls():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Input Data for Solar Panel Calculator"
    bold = Font(bold=True)
    ws["A1"].font = bold
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for row in ws["B3:B11"]:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    for row in ws["A3:B11"]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.column_dimensions["A"].width = 23
    ws.column_dimensions["B"].width = 30
    ws["A3"] = "Address"
    ws["A4"] = "City"
    ws["A5"] = "State"
    ws["A6"] = "Country"
    ws["A7"] = "Rooftop size, m2"
    ws["A8"] = "Annual electricity consumption, kWh"
    ws["A9"] = "Solar panel module type"
    ws["A10"] = "Fixing array type"
    ws["A11"] = "Azimuth of rooftop"
    ws[
        "A13"
    ] = "* Please check README.md file for detailed information about solar panel module type, fixing array type and azimuth."
    ws["A8"].alignment = Alignment(wrap_text=True)
    small_font = Font(size=9, italic=True)
    ws["A13"].font = small_font
    ws.merge_cells("A13:B14")
    ws["A13"].alignment = Alignment(wrap_text=True)
    module = '"Standard,Premium,Thin film"'
    array = '"Fixed(open track),Fixed(roof mount),1-Axis,1-Axis Backtracking,2-Axis"'
    azimuth = '"0,45,90,135,180,225,270,315"'
    # Inserting dropdown lists with data validation
    dv = DataValidation(type="list", formula1=module)
    dv.prompt = "Please select solar panel module type"
    dv.promptTitle = "Module type selection"
    dv.error = "Please choose from the list"
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(ws["B9"])
    ws["B9"].value = "Standard"
    dv.showInputMessage = True
    dv.showErrorMessage = True
    dv2 = DataValidation(type="list", formula1=array)
    dv2.prompt = "Please select fixing array type"
    dv2.promptTitle = "Fixing array type selection"
    dv2.error = "Please choose from the list"
    dv2.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv2)
    dv2.add(ws["B10"])
    ws["B10"].value = "Fixed(roof mount)"
    dv2.showInputMessage = True
    dv2.showErrorMessage = True
    dv3 = DataValidation(type="list", formula1=azimuth)
    dv3.prompt = "Please select azimuth"
    dv3.promptTitle = "Azimuth selection"
    dv3.error = "Please choose from the list"
    dv3.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv3)
    dv3.add(ws["B11"])
    ws["B11"].value = int("180")
    dv3.showInputMessage = True
    dv3.showErrorMessage = True
    dv4 = DataValidation(type="textLength", operator="lessThanOrEqual", formula1=15)
    dv4.prompt = "This field is required"
    dv4.promptTitle = "Mandatory field"
    ws.add_data_validation(dv4)
    dv4.add(ws["B4"])
    dv4.add(ws["B6"])
    dv4.add(ws["B7"])
    dv4.add(ws["B8"])
    dv4.showInputMessage = True
    wb.save("Data.xlsx")


if __name__ == "__main__":
    main()
