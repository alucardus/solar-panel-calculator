import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.chart import BarChart, Reference, Series


def main():
    pass

# Creating report file
def create_report(u, report):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Solar Panel Report"
    bold = Font(bold=True)
    ws["A1"].font = bold
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for row in ws["B3:B11"]:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    for row in ws["A3:C11"]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws["A3"] = "Annual electricity consumption"
    ws["A4"] = "Total solar panel power"
    ws["A5"] = "Total solar panel number"
    ws["A6"] = "Total square of solar panels"
    ws["A7"] = "Total solar energy annual output"
    ws["A8"] = "Total solar panel cost, including installation"
    ws["A8"].alignment = Alignment(wrap_text=True)
    ws["A9"] = "Utility price"
    ws["A10"] = "Total annual solar output cost"
    ws["A11"] = "Total repayment period"
    ws["B3"] = u.annual_consumption
    ws["B4"] = round(report[0], 1)
    ws["B5"] = int(report[2])
    ws["B6"] = round(report[1], 1)
    ws["B7"] = int(u.annual_output)
    ws["B8"] = int(report[3])
    ws["B9"] = round(u.price, 3)
    ws["B10"] = int(report[4])
    ws["B11"] = round(report[5], 1)
    ws["C3"] = "kWh"
    ws["C4"] = "kW"
    ws["C5"] = "unit"
    ws["C6"] = "m2"
    ws["C7"] = "kWh"
    ws["C8"] = "$"
    ws["C8"].alignment = Alignment(vertical="center")
    ws["C9"] = "$ per kWh"
    ws["C10"] = "$"
    ws["C11"] = "years"
    # Creating monthly output chart
    hide_font = Font(color="FFFFFF")
    months = list(calendar.month_name)
    for i, value in enumerate(months, start=1):
        ws.cell(row=i, column=4).value = value
        ws.cell(row=i, column=4).font = hide_font
    for i, value in enumerate(u.monthly_output, start=2):
        ws.cell(row=i, column=5).value = value
        ws.cell(row=i, column=5).font = hide_font
    chart = BarChart()
    labels = Reference(ws, min_col=4, min_row=2, max_row=13, max_col=4)
    data = Reference(ws, min_col=5, min_row=2, max_row=13)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    chart.title = "Monthly Solar Output"
    chart.y_axis.title = "kWh"
    chart.legend = None
    ws.add_chart(chart, "E1")
    wb.save("Report.xlsx")


if __name__ == "__main__":
    main()
